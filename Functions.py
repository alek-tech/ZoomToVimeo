from dotenv import load_dotenv
import vimeo
import os
import base64
import requests
from tqdm import tqdm
from time import sleep
from openpyxl import load_workbook
from vimeo.exceptions import UploadAttemptCreationFailure
from pathvalidate import sanitize_filepath, sanitize_filename


def load_completed_downloads_ids(self):
    try:
        with open(self.COMPLETED_DOWNLOADS_LOG, 'r') as fd:
            for line in fd:
                self.COMPLETED_DOWNLOADS_IDS.add(line.strip())
    except FileNotFoundError:
        print("\nLog file not found. Creating new log file: ",self.COMPLETED_DOWNLOADS_LOG)              
        with open(self.COMPLETED_DOWNLOADS_LOG, 'a') as log:
            log.write('\n')
            log.flush
        print()

        
def load_completed_uploads_ids(self):
    try:
        with open(self.COMPLETED_UPLOADS_LOG, 'r') as fd:
            for line in fd:
                self.COMPLETED_UPLOADS_IDS.add(line.strip())
    except FileNotFoundError:
        print("\nLog file not found. Creating new log file: ", self.COMPLETED_UPLOADS_LOG)
        with open(self.COMPLETED_UPLOADS_LOG, 'a') as log:
            log.write('\n')
            log.flush
        print()

        

# CONNECT TO VIMEO API
# POGLEDAT ČE JE KAKŠEN BOLŠI NAČIN ZA UPROABIT VIMEO.CLIENT IZ MAINA
def vimeo_connect():
    load_dotenv() 
    TOKEN = os.environ.get("vimeo_token")
    KEY = os.environ.get("vimeo_key")
    SECRET = os.environ.get("vimeo_secret")
    vimeo_connect.client = vimeo.VimeoClient(token=TOKEN, key=KEY, secret=SECRET)
    uri = "https://api.vimeo.com/tutorial"
    response = vimeo_connect.client.get(uri)
    if not response.ok:
        print(response)
        print("Could not connect to Vimeo API. Check credentials or URL.")
        exit(1)
    else:
        print("Vimeo API connection successful.")

# GET ZOOM S-S OAUTH-APP TOKEN (EXPIRES EVERY HOUR)
def get_token():
    try:
        load_dotenv() 
        ACCOUNT__ID = os.environ.get("zoom_account_id")
        SECRET = os.environ.get("zoom_password")
        CLIENT_ID = os.environ.get("zoom_user")

        url = "https://zoom.us/oauth/token?grant_type=account_credentials&account_id=" + ACCOUNT__ID
        auth = CLIENT_ID + ':' + SECRET
        encoded_auth = base64.b64encode(auth.encode()).decode()
        headers = {"Authorization" : "Basic " + encoded_auth}

        response = requests.request("POST", url, headers=headers) 
        token = response.json()["access_token"]

        print("Zoom Token recieved.\n")
        return token 
    except:
        print("Couldn't refresh ZOOM Token. Check get_token() function.\n")
        exit(1)

    
def get_all_users(self):
    print("\nGetting all users...\n")
    self.auth_header = {'Authorization': f'Bearer {self.zoom_token}'}
    params = {'page_size' : '300'}
    API_ENDPOINT_USERS = 'https://api.zoom.us/v2/users?status=active'
    try:
        response = requests.get(url=API_ENDPOINT_USERS, headers=self.auth_header, params=params)
    except:
        print("\n Couldn't get Users from Zoom, check credentials.")
        exit(1)

    json = response.json()
    # print(json)

    # GETS EMAILS OF ALL USERS
    n_of_pages = json['page_count']
    next_page = json['next_page_token']
    emails = []
    for i in range(n_of_pages):
        sleep(1.1)
        params = {'page_size' : '300',
                  'next_page_token' : next_page}
        response = requests.get(url=API_ENDPOINT_USERS,
                headers=self.auth_header, params = params)
        json = response.json()
        emails.extend(i['email'] for i in json['users'])
        next_page = json['next_page_token']
        #[print(i['email']) for i in json['users']]
    return emails

# GENERATOR FUNC THAT YIELDS DATES
# SINCE DATES_GEN YIELDS 2 ITEMS, WE CAN LOOP THROUGH IT LIKE FOR A,B IN DATES_GEN(START, END, DELTA)
# WE THEN MAKE A REQUEST AT EVERY ITERATION WITH THE NEW DATE GENERATED AS NEW PARAMS FOR THE REQUEST
def dates_gen(start, end, delta):
    current_date = start
    while current_date < end:
        yield current_date, min(current_date + delta, end)
        current_date += delta

# USES DATES FROM DATES_GEN FUNC, RETURNS PARAMS FOR NEXT REQUEST
def params(rec_start_date, rec_end_date, em):
    return {'userId': em,'page_size': '300','from': rec_start_date,'to': rec_end_date}

# EXTRACTS INFO ABOUT EVERY FILE FOUND IN CURRENT RECORDING INCLUDING DOWNLOAD URL
def get_recordings_files(self,rec):
    files_info = []
    recording_name = rec['topic']
    if rec.get('recording_files') == None:
      print("==> NO recording files here! Skipping")
      return files_info
    for download in rec['recording_files']:
        file_type = download['file_type']
        file_extension = download['file_extension']
        recording_id = download['id']
        recording_date = download['recording_start']
        if file_type == "":
            recording_type = 'incomplete'
        elif file_type != "TIMELINE":
            recording_type = download['recording_type']
        else:
            recording_type = download['file_type']

        #########################
        # TO FILTER SPECIFIC TYPE OF FILE UNCOMMENT NEXT 3 LINES, ADD FILE TYPE BETWEEN QUOTES AND COMMENT THE 2 LINES UNDER DEFINE DOWNLOAD URL
        # if recording_type == 'shared_screen_with_speaker_view':
        #     download_url = download['download_url'] + "?access_token=" + self.zoom_token
        #     files_info.append((file_type, file_extension, download_url, recording_type, recording_id, recording_name, recording_date))
        #########################    

        # DEFINE DOWNLOAD URL
        download_url = download['download_url'] + "?access_token=" + self.zoom_token
        files_info.append((file_type, file_extension, download_url, recording_type, recording_id, recording_name, recording_date))
        # else:
        #     pass
    return files_info

def download_file(self,download_url, em, filename):
    # REMOVE ODD CHARACTERS FROM RECORDING NAME
    # filename = re.sub(r'[^a-zA-Z0-9]', ' ', recording_name)
    # CREATE DWN PATH  
    try:
        full_foldername = os.sep.join([self.DOWNLOAD_DIRECTORY, em])
        self.correct_foldername = sanitize_filepath(full_foldername, platform = "auto")
        correct_filename = sanitize_filename(filename, platform = "auto")
        download_path = os.sep.join([self.correct_foldername, correct_filename])
        # print("\n",download_path,"\n")
        print("\nDownloading file ", correct_filename, "to ", self.correct_foldername + "\n")
        # CREATES FOLDER, DOESNT THROW ERROR IF FOLDER ALREADY EXISTS
        os.makedirs(self.correct_foldername, exist_ok=True)
        # DOWNLOAD FILE
        response = requests.get(download_url, stream=True)

        total_size = int(response.headers.get('content-length', 0))
        # DEFINE DOWNLOAD CHUNK SIZE
        block_size = 32 * 1024  # 32 Kibibytes
        # create TQDM progress bar
        t = tqdm(total=total_size, unit='iB', unit_scale=True)
        # WRITE FILE TO DWN PATH
        # try:
        with open(rf"{download_path}", 'wb') as fd:
            for chunk in response.iter_content(block_size):
                t.update(len(chunk))
                fd.write(chunk)  # write video chunk to disk
        t.close()
        print("File downloaded.")
        return True
    except Exception as e:
        print("There was an error downloading the recording, skipping to next one.")
        print(e)
        return False

def upper_inputs(word):
    word = word.split()
    word = [a.upper() for a in word ]
    word =  " ".join(word)
    return word

def start_program(start):
    check = False
    while check == False:
        cli = upper_inputs(input("Type 'ALL' or 'SINGLE' : "))
        if cli == "ALL":
            check = True
            start.all()
        elif cli == "SINGLE":
            check = True
            start.single()
        else:
            print("Wrong input, try again. Press ctr-l C to exit.")

def save_user_folder(em, folder_id):
    full_path = os.path.realpath(__file__)
    dir = os.path.dirname(full_path)
    saved_users = os.sep.join([dir, "All_users.xlsx"])
    wb = load_workbook(saved_users)
    page = wb.active
    user = em
    folder = folder_id
    page.append([user, folder])
    wb.save(saved_users)

def user_check(em):
    full_path = os.path.realpath(__file__)
    dir = os.path.dirname(full_path)
    saved_users = os.sep.join([dir, "All_users.xlsx"])
    wb = load_workbook(saved_users)
    page = wb.active
    check = False
    for cell in page['A']:
        if cell.value is not None:
            if em == cell.value:
                check = True
                folder_id = page.cell(row=cell.row,column=2)
            else:
                pass
        else:
            print("Remove cell ", cell, "from all_users.xlsx")
            exit(0)
    if check == True:
        return folder_id.value
    else:
        return False
    
def specific_folder():
    check = False
    while check == False:
        cli = upper_inputs(input("\nAny specific Vimeo folder to upload the recordings to? \nType 'YES' or 'NO': "))
        if cli == "YES":
            cli_2 = (input("Provide Folder id: "))
            check = True
            return cli_2
        elif cli == "NO":
            check = True
            return False
        else:
            print("Wrong input, try again. Press ctr-l C to exit.")
        
def choose_dates(self):
    check = False
    while check == False:
        cli = upper_inputs(input("Any starting dates to begin downloads from? \nType 'YES' or 'NO': "))
        if cli == "YES":
            check = True
            self.RECORDING_START_YEAR = int(input("YEAR: "))
            self.RECORDING_START_MONTH = int(input("MONTH: "))
            self.RECORDING_START_DAY = int(input("DAY: "))
        elif cli == "NO":
            check = True
            return False
        else:
            print("Wrong input, try again. Press ctr-l C to exit.")
        
        
def single_upload(self, em, filename, vimeo_folder_id, recording_id):

    if recording_id in self.COMPLETED_UPLOADS_IDS:
        print("==> Skipping already uploaded recording: ", recording_id)
        pass
    else:
        # VIMEO RESTRICTS UPLOAD OF FILES WITH NAME LENGHT OVER 128 CHAR
        if len(filename) > 128:
            filename = filename[:128]
        file_path = os.sep.join([self.correct_foldername, filename])
        sleep(1.1)
        # CREATE NEW FOLDER NAMED AFTER EMAIL AND UPLOAD FILE INSIDE IF FOLDER ID IS NOT SPECIFIED IN SINGLE_USERS.XLSX
        if vimeo_folder_id != vimeo_folder_id:
            if not user_check(em):
                print("No folder id")
                headers = {"Authorization": f"bearer {self.vimeo_token}",
                            "Content-Type": "application/json",
                            "Accept": "application/vnd.vimeo.*+json;version=3.4"}
                
                create_folder = vimeo_connect.client.post("https://api.vimeo.com/me/projects", headers=headers, data={'name':em})
                splitted_folder_id = create_folder.json()['uri'].split("/")
                vimeo_folder_id = splitted_folder_id[4]
                # SAVE USER FOLDER PAIR
                save_user_folder(em, vimeo_folder_id)
                print("New folder id:", vimeo_folder_id)
                # UPLOAD FILE
                try:
                    uri = vimeo_connect.client.upload(file_path, data={'name':filename,'folder_uri':'/folders/'+str(vimeo_folder_id),'privacy.embed':'whitelist'})# print("PARAMS:", file_name, str(file), '/folders/'+str(folder_id), 'whitelist' )
                    # print(uri)
                    # print(uri.content)
                    # print(uri.status_code)
                except UploadAttemptCreationFailure:
                    print("Upload error, Check folder id or request params.")
                    exit(0)
            else:
                # IF USER FOLDER FOUND IN All_users.xlsx WILL UPLOAD FILE IN THAT FOLDER
                vimeo_folder_id = int(user_check(em))
                # UPLOAD FILE
                try:
                    uri = vimeo_connect.client.upload(file_path, data={'name':filename,'folder_uri':'/folders/'+str(vimeo_folder_id),'privacy.embed':'whitelist'})
                    # print(uri)
                    # print(uri.content)
                    # print(uri.status_code)
                except UploadAttemptCreationFailure:
                    print("Upload error, Check folder id or request params.")
                    exit(0)
        else:
            print("Current folder id: ", vimeo_folder_id)
            # UPLOAD FILE
            try:
                uri = vimeo_connect.client.upload(file_path, data={'name':filename,'folder_uri':'/folders/'+str(vimeo_folder_id),'privacy.embed':'whitelist'})# print("PARAMS:", file_name, str(file), '/folders/'+str(folder_id), 'whitelist' )
                # print(uri)
                # print(uri.content)
                # print(uri.status_code)
            except UploadAttemptCreationFailure:
                print("Upload error, Check folder id or request params.")
                exit(0)

        print("\nVideo",filename,"uploaded to Vimeo succesfully!\n") 
        print ('Your video URI is:',uri,"\n")
        # if successful, write the ID of this recording to the completed uploads ids
        with open(self.COMPLETED_UPLOADS_LOG, 'a') as log:
            self.COMPLETED_UPLOADS_IDS.add(recording_id)
            log.write(recording_id)
            log.write('\n')
            log.flush()
        sleep(1)


def all_upload(self, em, filename, vimeo_folder_id, recording_id):

    if recording_id in self.COMPLETED_UPLOADS_IDS:
        print("==> Skipping already uploaded recording: ", recording_id)
        pass
    else:
        # VIMEO RESTRICTS UPLOAD OF FILES WITH NAME LENGHT OVER 128 CHAR
        if len(filename) > 128:
            filename = filename[:128]
        file_path = os.sep.join([self.correct_foldername, filename])
        sleep(1.1)
        # CREATE NEW FOLDER FOR EVERY NEW USER AND PUT HIS RECORDINGS INSIDE.
        if vimeo_folder_id == None:
            print("No folder id")
            headers = {"Authorization": f"bearer {self.vimeo_token}",
                        "Content-Type": "application/json",
                        "Accept": "application/vnd.vimeo.*+json;version=3.4"}
            
            create_folder = vimeo_connect.client.post("https://api.vimeo.com/me/projects", headers=headers, data={'name':em})
            splitted_folder_id = create_folder.json()['uri'].split("/")
            vimeo_folder_id = splitted_folder_id[4]
            save_user_folder(em, vimeo_folder_id)
            print("New folder id:", vimeo_folder_id)
            # UPLOAD FILE
            try:
                uri = vimeo_connect.client.upload(file_path, data={'name':filename,'folder_uri':'/folders/'+str(vimeo_folder_id),'privacy.embed':'whitelist'})
            except UploadAttemptCreationFailure:
                print("Upload error, Check folder id or request params.")
                exit(0)
        else:
            # UPLOAD FILE
            try:
                uri = vimeo_connect.client.upload(file_path, data={'name':filename,'folder_uri':'/folders/'+str(vimeo_folder_id),'privacy.embed':'whitelist'})
            except UploadAttemptCreationFailure:
                print("Upload error, Check folder id or request params.")
                exit(0)

        print("\nVideo",filename,"uploaded to Vimeo succesfully!\n") 
        print ('Your video URI is:',uri,"\n")
        # if successful, write the ID of this recording to the completed uploads ids
        with open(self.COMPLETED_UPLOADS_LOG, 'a') as log:
            self.COMPLETED_UPLOADS_IDS.add(recording_id)
            log.write(recording_id)
            log.write('\n')
            log.flush()
        # sleep(1)
